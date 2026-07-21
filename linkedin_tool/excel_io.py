# -*- coding: utf-8 -*-
"""
excel_io.py — In- en uitlezen van Excel-bestanden (via openpyxl, geen pandas).

Twee taken:
1. `lees_namen`  : leest namen.xlsx  (kolom A = naam, kolom B = context).
2. `schrijf_kandidaten` : schrijft kandidaten.xlsx voor de batch-matching.
   Deze functie wordt na ELKE naam aangeroepen (tussenstand), zodat je bij een
   onderbreking niets kwijtraakt.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook


def lees_namen(pad):
    """Lees namen.xlsx en geef een lijst (naam, context)-paren terug.

    - Kolom A: naam (verplicht).
    - Kolom B: context (optioneel, bv. organisatie of regio).
    - Een eventuele koprij wordt automatisch overgeslagen als in A1 letterlijk
      "naam" staat.
    - Lege rijen worden overgeslagen.
    """
    bestand = Path(pad)
    if not bestand.exists():
        raise FileNotFoundError(
            f"Bestand niet gevonden: {bestand}. Maak een Excel met in kolom A de "
            f"namen en (optioneel) in kolom B context."
        )

    werkboek = load_workbook(bestand, read_only=True, data_only=True)
    blad = werkboek.active

    namen = []
    for rij_nr, rij in enumerate(blad.iter_rows(min_row=1, max_col=2, values_only=True), start=1):
        naam = (rij[0] or "")
        context = (rij[1] or "") if len(rij) > 1 else ""

        naam = str(naam).strip()
        context = str(context).strip()

        # Koprij overslaan (als iemand een kop "naam" heeft gezet).
        if rij_nr == 1 and naam.lower() in ("naam", "name"):
            continue

        # Lege rijen overslaan.
        if not naam:
            continue

        namen.append((naam, context))

    werkboek.close()
    return namen


def schrijf_kandidaten(resultaten, pad):
    """Schrijf de matching-resultaten naar kandidaten.xlsx.

    `resultaten` is een lijst met dicts, bv.:
        {
          "naam": "Jan Jansen",
          "context": "Utrecht",
          "kandidaten": [ {"omschrijving": "...", "profiel_url": "..."}, ... ]
        }

    Layout: per gezochte naam één blok met daaronder de kandidaten. Overzichtelijk
    en makkelijk door te lezen in Excel.

    We schrijven het bestand steeds volledig opnieuw. Zo is elke tussenstand een
    compleet, bruikbaar bestand.
    """
    werkboek = Workbook()
    blad = werkboek.active
    blad.title = "Kandidaten"

    # Koprij.
    blad.append(["Gezochte naam", "Context", "Kandidaat (omschrijving)", "Profiel-URL"])

    for item in resultaten:
        naam = item.get("naam", "")
        context = item.get("context", "")
        kandidaten = item.get("kandidaten", [])

        if not kandidaten:
            # Toch een rij, zodat je ziet dat er niets gevonden is.
            blad.append([naam, context, "(geen kandidaten gevonden)", ""])
            continue

        for kandidaat in kandidaten:
            blad.append([
                naam,
                context,
                kandidaat.get("omschrijving", ""),
                kandidaat.get("profiel_url", ""),
            ])

    # Kolombreedtes een beetje instellen voor de leesbaarheid.
    breedtes = {"A": 22, "B": 18, "C": 55, "D": 45}
    for kolom, breedte in breedtes.items():
        blad.column_dimensions[kolom].width = breedte

    werkboek.save(pad)
